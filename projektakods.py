from openpyxl import workbook, load_workbook
wb=load_workbook("Studenta_kalendars.xlsx")
ws=wb.active
max_row=ws.max_row
print(max_row)
total_majasdarbu_stundas=0
total_sporta_stundas=0
total_studiju_stundas=0
total_darba_stundas=0
total=0
for row in range(4,max_row+1):
    studiju_minutes = ws['C' + str(row)].value
    darba_stundas = ws['F' + str(row)].value
    if studiju_minutes is not None:
        studiju_stundas=studiju_minutes/60
        total_studiju_stundas+=studiju_stundas
        total_darba_stundas+=darba_stundas
        total+=studiju_stundas+darba_stundas
sporta_stundas=float(input("Ievadīt cik stundas pavadijāt sportojot: "))
total_sporta_stundas=sporta_stundas
majasdarbi_stundas=float(input("Ievadīt cik stundas pavadijāt pildot majasdarbus: "))
total_majasdarbu_stundas=majasdarbi_stundas
total_produktivitates_stundas = total_studiju_stundas + total_darba_stundas + total_sporta_stundas+total_majasdarbu_stundas
produktivitates_studiju_procenti = (total_studiju_stundas / total_produktivitates_stundas) * 100
produktivitates_darba_procenti = (total_darba_stundas / total_produktivitates_stundas) * 100
produktivitates_sporta_procenti = (total_sporta_stundas / total_produktivitates_stundas) * 100
produktivitates_majasdarbu_procenti = (total_majasdarbu_stundas / total_produktivitates_stundas) * 100

print(f"kopa Studiju Stundas: {total_studiju_stundas} stundas")
print(f"kopa Darba Stundas: {total_darba_stundas} stundas")
print(f"kopa Sporta Stundas: {total_sporta_stundas} stundas")
print(f"kopa MD Stundas: {total_majasdarbu_stundas} stundas")
print(f"kopa Produktivitate Stundas: {total_produktivitates_stundas} stundas")
print(f"Produktivitate studiju Procenti: {produktivitates_studiju_procenti:.2f}%")
print(f"Produktivitate darba Procenti: {produktivitates_darba_procenti:.2f}%")
print(f"Produktivitate sporta Procenti: {produktivitates_sporta_procenti:.2f}%")
print(f"Produktivitate majasdarbu Procenti: {produktivitates_majasdarbu_procenti:.2f}%")
